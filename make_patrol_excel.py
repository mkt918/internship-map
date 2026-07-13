import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

SRC = r'C:\antigravity\public\00_作業\2026-06\インターンシップ管理\patrol_routes.json'
OUT = r'C:\antigravity\public\00_作業\2026-06\インターンシップ管理\教員見回り担当表.xlsx'

with open(SRC, encoding='utf-8') as f:
    data = json.load(f)

TEACHER_COLORS = ['D94F4F','3A7EC8','27AE60','E67E22','8E44AD']
HEADER_BG = '1A3A5C'
INPUT_BG  = 'FFF9C4'
WHITE     = 'FFFFFF'
ALT_BG    = 'F0F4FF'

def side(style='thin', color='CCCCCC'):
    return Side(style=style, color=color)

def border(all='thin'):
    s = side(all)
    return Border(left=s, right=s, top=s, bottom=s)

def fill(hex_color):
    return PatternFill('solid', start_color=hex_color, fgColor=hex_color)

def font(bold=False, size=10, color='000000', name='Meiryo'):
    return Font(bold=bold, size=size, color=color, name=name)

wb = Workbook()

ws_all = wb.active
ws_all.title = '全日程一覧'

ws_all.merge_cells('A1:G1')
ws_all['A1'] = 'インターンシップ 教員見回り担当表'
ws_all['A1'].font = Font(bold=True, size=14, color=WHITE, name='Meiryo')
ws_all['A1'].fill = fill(HEADER_BG)
ws_all['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws_all.row_dimensions[1].height = 28

col_widths = [12, 18, 18, 20, 38, 12, 16]
for i, w in enumerate(col_widths, 1):
    ws_all.column_dimensions[get_column_letter(i)].width = w

headers = ['日付', '教員番号', '担当教員名', '業種', '訪問企業（順番）', '生徒数', '移動距離']
for c, h in enumerate(headers, 1):
    cell = ws_all.cell(row=2, column=c, value=h)
    cell.font = Font(bold=True, size=10, color=WHITE, name='Meiryo')
    cell.fill = fill('2D5A8E')
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = border()
ws_all.row_dimensions[2].height = 22

r = 3
for day in data:
    date_str = day['date']
    start_r = r
    for route in day['routes']:
        t_idx = route['teacher'] - 1
        tc = TEACHER_COLORS[t_idx % len(TEACHER_COLORS)]
        stops_str = ' → '.join(s['name'] for s in route['stops'])
        cats = list(dict.fromkeys(cat.split(',')[0].strip() for s in route['stops'] for cat in [s['cat']]))
        cat_str = '・'.join(cats)

        row_bg = WHITE if (r - 3) % 2 == 0 else ALT_BG

        ws_all.cell(row=r, column=1, value=date_str)

        cell_b = ws_all.cell(row=r, column=2, value=f'教員 {route["teacher"]}')
        cell_b.font = Font(bold=True, size=10, color=WHITE, name='Meiryo')
        cell_b.fill = fill(tc)
        cell_b.alignment = Alignment(horizontal='center', vertical='center')
        cell_b.border = border()

        cell_c = ws_all.cell(row=r, column=3, value='')
        cell_c.fill = fill(INPUT_BG)
        cell_c.border = Border(
            left=Side(style='medium', color='F0C000'),
            right=Side(style='medium', color='F0C000'),
            top=Side(style='thin', color='F0C000'),
            bottom=Side(style='thin', color='F0C000'),
        )
        cell_c.alignment = Alignment(horizontal='center', vertical='center')

        ws_all.cell(row=r, column=4, value=cat_str).border = border()
        ws_all.cell(row=r, column=4).alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        ws_all.cell(row=r, column=4).fill = fill(row_bg)

        cell_e = ws_all.cell(row=r, column=5, value=stops_str)
        cell_e.border = border()
        cell_e.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        cell_e.fill = fill(row_bg)

        ws_all.cell(row=r, column=6, value=route['n_students']).border = border()
        ws_all.cell(row=r, column=6).alignment = Alignment(horizontal='center', vertical='center')
        ws_all.cell(row=r, column=6).fill = fill(row_bg)

        ws_all.cell(row=r, column=7, value=f'{route["total_km"]}km').border = border()
        ws_all.cell(row=r, column=7).alignment = Alignment(horizontal='center', vertical='center')
        ws_all.cell(row=r, column=7).fill = fill(row_bg)

        ws_all.row_dimensions[r].height = 32
        r += 1

    end_r = r - 1
    if end_r > start_r:
        ws_all.merge_cells(f'A{start_r}:A{end_r}')
    cell_a = ws_all.cell(row=start_r, column=1)
    cell_a.font = Font(bold=True, size=10, name='Meiryo', color='1A3A5C')
    cell_a.fill = fill('D6E4F0')
    cell_a.alignment = Alignment(horizontal='center', vertical='center')
    cell_a.border = Border(
        left=Side(style='medium', color='1A3A5C'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='medium', color='1A3A5C'),
        bottom=Side(style='medium', color='1A3A5C'),
    )

for day in data:
    date_str = day['date']
    safe = date_str.replace('/', '-').replace('(', '').replace(')', '')
    ws = wb.create_sheet(title=safe)

    ws.merge_cells('A1:F1')
    ws['A1'] = f'見回り担当表　{date_str}　({day["n_companies"]}社 / {day["n_students"]}名)'
    ws['A1'].font = Font(bold=True, size=13, color=WHITE, name='Meiryo')
    ws['A1'].fill = fill(HEADER_BG)
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 26

    col_widths2 = [16, 20, 14, 38, 12, 14]
    for i, w in enumerate(col_widths2, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    headers2 = ['担当教員名', '業種', '訪問順', '訪問先企業', '生徒数', '移動距離']
    for c, h in enumerate(headers2, 1):
        cell = ws.cell(row=2, column=c, value=h)
        cell.font = Font(bold=True, size=10, color=WHITE, name='Meiryo')
        cell.fill = fill('2D5A8E')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border()
    ws.row_dimensions[2].height = 20

    r2 = 3
    for route in day['routes']:
        t_idx = route['teacher'] - 1
        tc = TEACHER_COLORS[t_idx % len(TEACHER_COLORS)]
        start_r2 = r2

        for si, stop in enumerate(route['stops']):
            row_bg2 = WHITE if si % 2 == 0 else 'F8F8F8'

            cell_a2 = ws.cell(row=r2, column=1, value='')
            cell_a2.fill = fill(INPUT_BG)
            cell_a2.border = Border(
                left=Side(style='medium', color='F0C000'),
                right=Side(style='medium', color='F0C000'),
                top=Side(style='thin' if si > 0 else 'medium', color='F0C000'),
                bottom=Side(style='thin' if si < len(route['stops'])-1 else 'medium', color='F0C000'),
            )
            cell_a2.alignment = Alignment(horizontal='center', vertical='center')

            cats2 = [c.strip() for c in stop['cat'].split(',')]
            ws.cell(row=r2, column=2, value=cats2[0]).border = border()
            ws.cell(row=r2, column=2).fill = fill(row_bg2)
            ws.cell(row=r2, column=2).alignment = Alignment(horizontal='left', vertical='center')

            cell_c2 = ws.cell(row=r2, column=3, value=f'  {si+1}社目')
            cell_c2.font = Font(bold=True, size=10, color=WHITE, name='Meiryo')
            cell_c2.fill = fill(tc)
            cell_c2.alignment = Alignment(horizontal='center', vertical='center')
            cell_c2.border = border()

            stu = '・'.join(stop['students'])
            ws.cell(row=r2, column=4, value=f'{stop["name"]}　　[生徒: {stu}]').border = border()
            ws.cell(row=r2, column=4).alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            ws.cell(row=r2, column=4).fill = fill(row_bg2)

            ws.cell(row=r2, column=5, value=len(stop['students'])).border = border()
            ws.cell(row=r2, column=5).alignment = Alignment(horizontal='center', vertical='center')
            ws.cell(row=r2, column=5).fill = fill(row_bg2)

            label = '学校から' if si == 0 else '前より'
            ws.cell(row=r2, column=6, value=f'{label} {stop["dist_from_prev"]}km').border = border()
            ws.cell(row=r2, column=6).alignment = Alignment(horizontal='center', vertical='center')
            ws.cell(row=r2, column=6).fill = fill(row_bg2)

            ws.row_dimensions[r2].height = 28
            r2 += 1

        end_r2 = r2 - 1
        if end_r2 > start_r2:
            ws.merge_cells(f'A{start_r2}:A{end_r2}')

        ws.cell(row=r2, column=1, value='合計')
        ws.cell(row=r2, column=1).font = Font(bold=True, size=9, color='666666', name='Meiryo')
        ws.cell(row=r2, column=1).fill = fill('EEEEEE')
        ws.cell(row=r2, column=1).border = border()
        ws.cell(row=r2, column=1).alignment = Alignment(horizontal='center')

        ws.merge_cells(f'B{r2}:C{r2}')
        ws.cell(row=r2, column=2, value=f'教員 {route["teacher"]}　計 {route["n_companies"]}社 / {route["n_students"]}名')
        ws.cell(row=r2, column=2).font = Font(bold=True, size=9, color='333333', name='Meiryo')
        ws.cell(row=r2, column=2).fill = fill('EEEEEE')
        ws.cell(row=r2, column=2).border = border()
        ws.cell(row=r2, column=2).alignment = Alignment(horizontal='left')

        ws.merge_cells(f'D{r2}:E{r2}')
        ws.cell(row=r2, column=4, value='')
        ws.cell(row=r2, column=4).fill = fill('EEEEEE')
        ws.cell(row=r2, column=4).border = border()

        ws.cell(row=r2, column=6, value=f'計 {route["total_km"]}km')
        ws.cell(row=r2, column=6).font = Font(bold=True, size=9, color='333333', name='Meiryo')
        ws.cell(row=r2, column=6).fill = fill('EEEEEE')
        ws.cell(row=r2, column=6).border = border()
        ws.cell(row=r2, column=6).alignment = Alignment(horizontal='center')
        ws.row_dimensions[r2].height = 18
        r2 += 2

wb.save(OUT)
print(f'保存: {OUT}')
