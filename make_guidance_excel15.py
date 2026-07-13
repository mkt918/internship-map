import json, csv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

BASE = r'C:\antigravity\public\00_作業\2026-06\インターンシップ管理'
with open(BASE + r'\teacher_identity.json', encoding='utf-8') as f:
    teachers = json.load(f)
with open(BASE + r'\classroom_plan15.json', encoding='utf-8') as f:
    plan = json.load(f)

comp_addr = {}
with open(BASE + r'\企業マスターDB.csv', encoding='utf-8-sig') as f:
    for r in csv.DictReader(f):
        comp_addr[r['企業名'].strip()] = r.get('住所', '').strip()

OUT = BASE + r'\7-17事前指導_教室割当.xlsx'

HDR_BG   = '1A3A5C'
INPUT_BG = 'FFF9C4'
WHITE    = 'FFFFFF'
LIGHT    = 'F0F4FA'
PAIR_COLORS = ['C0392B', '2471A3']

def fill(c): return PatternFill('solid', fgColor=c)
def fnt(bold=False, size=10, color='000000', name='Meiryo'):
    return Font(bold=bold, size=size, color=color, name=name)
def side(s='thin', c='BBBBBB'): return Side(style=s, color=c)
def bdr(t='thin'): s=side(t); return Border(left=s,right=s,top=s,bottom=s)
def gold_bdr():
    m = Side(style='medium', color='E0A800')
    return Border(left=m, right=m, top=m, bottom=m)

wb = Workbook()

ws = wb.active
ws.title = '7-17教室割当一覧'
ws.column_dimensions['A'].width = 9
ws.column_dimensions['B'].width = 9
ws.column_dimensions['C'].width = 16
ws.column_dimensions['D'].width = 10
ws.column_dimensions['E'].width = 40
ws.column_dimensions['F'].width = 38
ws.column_dimensions['G'].width = 9
ws.column_dimensions['H'].width = 9

ws.row_dimensions[1].height = 30
ws.merge_cells('A1:H1')
ws['A1'] = '7月17日(金) 4限　インターンシップ事前指導　教室割当表（教員16名体制）'
ws['A1'].font = fnt(True, 14, WHITE)
ws['A1'].fill = fill(HDR_BG)
ws['A1'].alignment = Alignment(horizontal='center', vertical='center')

headers = ['教室', '教員\nNo.', '担当教員名\n（ここに記入）', '担当日数', '巡回予定日と訪問先', '生徒（年-組-番）', '生徒数', '教室計']
ws.row_dimensions[2].height = 26
for c, h in enumerate(headers, 1):
    cell = ws.cell(row=2, column=c, value=h)
    cell.font = fnt(True, 9, WHITE)
    cell.fill = fill('2D5A8E')
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = bdr()

r = 3
name_row_map = {}
for room in plan:
    start_r = r
    for pi, member in enumerate(room['teachers']):
        tid = member['teacher']
        routes = member['routes']
        total_stu = member['total_students']
        all_students = [s for rt in routes for s in rt['students']]
        pc = PAIR_COLORS[pi % len(PAIR_COLORS)]
        row_bg = WHITE if pi == 0 else LIGHT

        ws.cell(row=r, column=1)

        cb = ws.cell(row=r, column=2, value=f'教員{tid}')
        cb.font = fnt(True, 9, WHITE); cb.fill = fill(pc)
        cb.alignment = Alignment(horizontal='center', vertical='center')
        cb.border = bdr()

        cc = ws.cell(row=r, column=3, value='')
        cc.fill = fill(INPUT_BG); cc.border = gold_bdr()
        cc.alignment = Alignment(horizontal='center', vertical='center')
        name_row_map[tid] = r

        cd = ws.cell(row=r, column=4, value=f'{len(routes)}日')
        cd.border = bdr(); cd.fill = fill(row_bg)
        cd.alignment = Alignment(horizontal='center', vertical='center')
        cd.font = fnt(False, 9)

        sched_lines = [f'{rt["date"]}: ' + ' → '.join(rt['companies']) for rt in routes]
        ce = ws.cell(row=r, column=5, value='\n'.join(sched_lines))
        ce.border = bdr(); ce.fill = fill(row_bg)
        ce.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        ce.font = fnt(False, 8)

        cf = ws.cell(row=r, column=6, value='　'.join(all_students))
        cf.border = bdr(); cf.fill = fill(row_bg)
        cf.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        cf.font = fnt(False, 8, '333333')

        cg = ws.cell(row=r, column=7, value=total_stu)
        cg.border = bdr(); cg.fill = fill(row_bg)
        cg.alignment = Alignment(horizontal='center', vertical='center')
        cg.font = fnt(True, 10)

        ws.row_dimensions[r].height = max(50, 16 * len(routes))
        r += 1

    end_r = r - 1
    if end_r > start_r:
        ws.merge_cells(f'A{start_r}:A{end_r}')
        ws.merge_cells(f'H{start_r}:H{end_r}')
    ca = ws.cell(row=start_r, column=1, value=f'教室{room["room"]}')
    ca.font = fnt(True, 12, WHITE); ca.fill = fill('333333')
    ca.alignment = Alignment(horizontal='center', vertical='center')
    ca.border = bdr('medium')

    ch = ws.cell(row=start_r, column=8, value=f'計\n{room["total_students"]}名')
    ch.font = fnt(True, 10, '1A3A5C'); ch.fill = fill('D6E4F0')
    ch.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ch.border = bdr('medium')

for room in plan:
    for member in room['teachers']:
        tid = member['teacher']
        routes = member['routes']
        pc = PAIR_COLORS[room['teachers'].index(member) % len(PAIR_COLORS)]
        name_ref = f"='7-17教室割当一覧'!C{name_row_map[tid]}"

        ws2 = wb.create_sheet(title=f'教員{tid}_詳細')
        ws2.page_setup.paperSize = ws2.PAPERSIZE_A4
        ws2.page_setup.orientation = 'portrait'
        ws2.column_dimensions['A'].width = 14
        ws2.column_dimensions['B'].width = 30
        ws2.column_dimensions['C'].width = 34
        ws2.column_dimensions['D'].width = 10

        ws2.row_dimensions[1].height = 32
        ws2.merge_cells('A1:D1')
        ws2['A1'] = f'教室{room["room"]}　事前指導（7/17 4限）'
        ws2['A1'].font = fnt(True, 14, WHITE)
        ws2['A1'].fill = fill(HDR_BG)
        ws2['A1'].alignment = Alignment(horizontal='center', vertical='center')

        ws2.row_dimensions[2].height = 26
        ws2['A2'] = '担当教員'
        ws2['A2'].font = fnt(True, 10, '333333')
        ws2['A2'].fill = fill('EEEEEE'); ws2['A2'].border = bdr()
        ws2['A2'].alignment = Alignment(horizontal='center', vertical='center')
        ws2.merge_cells('B2:D2')
        ws2['B2'] = name_ref
        ws2['B2'].fill = fill(INPUT_BG); ws2['B2'].border = gold_bdr()
        ws2['B2'].font = fnt(True, 12, '1A3A5C')
        ws2['B2'].alignment = Alignment(horizontal='left', vertical='center')

        total_stu = member['total_students']
        ws2.row_dimensions[3].height = 18
        ws2.merge_cells('A3:D3')
        ws2['A3'] = f'夏休み中の巡回予定：{len(routes)}日間 ／ 対象生徒 計{total_stu}名（教室{room["room"]}で他1名と同室）'
        ws2['A3'].font = fnt(True, 9, WHITE); ws2['A3'].fill = fill(pc)
        ws2['A3'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        ws2.row_dimensions[4].height = 6

        headers2 = ['巡回日', '訪問先企業', '担当生徒', '距離']
        ws2.row_dimensions[5].height = 18
        for c, h in enumerate(headers2, 1):
            cell = ws2.cell(row=5, column=c, value=h)
            cell.font = fnt(True, 9, WHITE); cell.fill = fill('2D5A8E')
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = bdr()

        r2 = 6
        for rt in routes:
            start = r2
            for ci, comp in enumerate(rt['companies']):
                addr = comp_addr.get(comp, '')
                row_bg = WHITE if ci % 2 == 0 else LIGHT
                ws2.row_dimensions[r2].height = 26
                c_comp = ws2.cell(row=r2, column=2, value=f'{comp}\n{addr}')
                c_comp.font = fnt(False, 8); c_comp.fill = fill(row_bg)
                c_comp.border = bdr(); c_comp.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                r2 += 1
            end = r2 - 1
            if end > start:
                ws2.merge_cells(f'A{start}:A{end}')
                ws2.merge_cells(f'D{start}:D{end}')
                ws2.merge_cells(f'C{start}:C{end}')
            cdate = ws2.cell(row=start, column=1, value=rt['date'])
            cdate.font = fnt(True, 10, WHITE); cdate.fill = fill(pc)
            cdate.alignment = Alignment(horizontal='center', vertical='center')
            cdate.border = bdr('medium')

            cdist = ws2.cell(row=start, column=4, value=f'{rt["total_km"]}km')
            cdist.font = fnt(False, 9); cdist.border = bdr()
            cdist.alignment = Alignment(horizontal='center', vertical='center')

            stu_str = '　'.join(rt['students'])
            cstu2 = ws2.cell(row=start, column=3, value=stu_str)
            cstu2.font = fnt(False, 8, '333333'); cstu2.fill = fill(WHITE)
            cstu2.border = bdr(); cstu2.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

        ws2.row_dimensions[r2].height = 16
        ws2.merge_cells(f'A{r2}:D{r2}')
        ws2.cell(row=r2, column=1, value='備考・メモ')
        ws2.cell(row=r2, column=1).font = fnt(True, 9, WHITE)
        ws2.cell(row=r2, column=1).fill = fill('7F8C8D')
        ws2.cell(row=r2, column=1).border = bdr()
        r2 += 1
        for _ in range(4):
            ws2.row_dimensions[r2].height = 18
            ws2.merge_cells(f'A{r2}:D{r2}')
            ws2.cell(row=r2, column=1, value='').fill = fill(WHITE)
            ws2.cell(row=r2, column=1).border = bdr('thin')
            r2 += 1

wb.save(OUT)
print(f'完了: {OUT}')
