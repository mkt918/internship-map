import json, csv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

BASE = r'C:\antigravity\public\00_作業\2026-06\インターンシップ管理'
SRC_JSON = BASE + r'\patrol_routes.json'
SRC_COMP = BASE + r'\企業マスターDB.csv'
OUT      = BASE + r'\教員見回り配布用.xlsx'

with open(SRC_JSON, encoding='utf-8') as f:
    data = json.load(f)

comp_addr = {}
with open(SRC_COMP, encoding='utf-8-sig') as f:
    for r in csv.DictReader(f):
        comp_addr[r['企業名'].strip()] = r.get('住所', '').strip()

TEACHER_COLORS = ['C0392B','2471A3','1E8449','D68910','76448A']
HDR_BG   = '1A3A5C'
INPUT_BG = 'FFF9C4'
WHITE    = 'FFFFFF'
LIGHT    = 'F0F4FA'

def fill(c): return PatternFill('solid', fgColor=c)
def fnt(bold=False, size=10, color='000000', name='Meiryo'):
    return Font(bold=bold, size=size, color=color, name=name)
def side(s='thin', c='BBBBBB'): return Side(style=s, color=c)
def bdr(t='thin'): s=side(t); return Border(left=s,right=s,top=s,bottom=s)
def gold_bdr():
    m=Side(style='medium',color='E0A800'); t=Side(style='thin',color='E0A800')
    return Border(left=m,right=m,top=t,bottom=t)

wb = Workbook()

ws_all = wb.active
ws_all.title = INDEX_SHEET = '全日程一覧'
ws_all.column_dimensions['A'].width = 11
ws_all.column_dimensions['B'].width = 13
ws_all.column_dimensions['C'].width = 22
ws_all.column_dimensions['D'].width = 34
ws_all.column_dimensions['E'].width = 32
ws_all.column_dimensions['F'].width = 28
ws_all.column_dimensions['G'].width = 12

ws_all.row_dimensions[1].height = 30
ws_all.merge_cells('A1:G1')
ws_all['A1'] = 'インターンシップ 教員見回り担当表'
ws_all['A1'].font = fnt(True, 14, WHITE)
ws_all['A1'].fill = fill(HDR_BG)
ws_all['A1'].alignment = Alignment(horizontal='center', vertical='center')

ws_all.row_dimensions[2].height = 24
for c, h in enumerate(['日付','教員番号','担当教員名 ▼ここに入力','訪問先企業','住　所','担当生徒（年-組-番）','距離'], 1):
    cell = ws_all.cell(row=2, column=c, value=h)
    cell.font = fnt(True, 9, WHITE)
    cell.fill = fill('2D5A8E')
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = bdr()

route_row_map = {}
r = 3
for day in data:
    start_r = r
    for route in day['routes']:
        tc = TEACHER_COLORS[(route['teacher']-1) % len(TEACHER_COLORS)]
        stops_str = ' → '.join(s['name'] for s in route['stops'])
        addr_str  = ' ／ '.join(comp_addr.get(s['name'], '') for s in route['stops'])
        all_stu   = [stu for s in route['stops'] for stu in s['students']]
        stu_str   = '　'.join(all_stu)

        safe_date = day['date'].replace('/','-').replace('(','').replace(')','')
        sheet_name = f'{safe_date}_教員{route["teacher"]}'
        route_row_map[sheet_name] = r

        ws_all.row_dimensions[r].height = 38
        row_bg = WHITE if (r-3) % 2 == 0 else LIGHT

        ws_all.cell(row=r, column=1)

        cb = ws_all.cell(row=r, column=2, value=f'教員 {route["teacher"]}')
        cb.font = fnt(True, 10, WHITE); cb.fill = fill(tc)
        cb.alignment = Alignment(horizontal='center', vertical='center')
        cb.border = bdr()

        cc = ws_all.cell(row=r, column=3, value='')
        cc.fill = fill(INPUT_BG); cc.border = gold_bdr()
        cc.alignment = Alignment(horizontal='left', vertical='center')
        cc.font = fnt(False, 11, '333333')

        cd = ws_all.cell(row=r, column=4, value=stops_str)
        cd.fill = fill(row_bg); cd.border = bdr()
        cd.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        cd.font = fnt(False, 9)

        ce = ws_all.cell(row=r, column=5, value=addr_str)
        ce.fill = fill(row_bg); ce.border = bdr()
        ce.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        ce.font = fnt(False, 8, '555555')

        cf = ws_all.cell(row=r, column=6, value=stu_str)
        cf.fill = fill(row_bg); cf.border = bdr()
        cf.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        cf.font = fnt(False, 8, '333333')

        cg = ws_all.cell(row=r, column=7, value=f'{route["total_km"]}km')
        cg.fill = fill(row_bg); cg.border = bdr()
        cg.alignment = Alignment(horizontal='center', vertical='center')
        cg.font = fnt(False, 9)

        r += 1

    end_r = r - 1
    if end_r > start_r:
        ws_all.merge_cells(f'A{start_r}:A{end_r}')
    ca = ws_all.cell(row=start_r, column=1, value=day['date'])
    ca.font = fnt(True, 10, '1A3A5C'); ca.fill = fill('D6E4F0')
    ca.alignment = Alignment(horizontal='center', vertical='center')
    ca.border = Border(left=Side(style='medium',color='1A3A5C'),
                       right=Side(style='thin',color='CCCCCC'),
                       top=Side(style='medium',color='1A3A5C'),
                       bottom=Side(style='medium',color='1A3A5C'))

import re
def dsort(d):
    m=re.search(r'(\d+)/(\d+)',d); return (int(m.group(1)),int(m.group(2))) if m else (99,99)

for day in sorted(data, key=lambda d: dsort(d['date'])):
    for route in day['routes']:
        tc = TEACHER_COLORS[(route['teacher']-1) % len(TEACHER_COLORS)]
        safe_date = day['date'].replace('/','-').replace('(','').replace(')','')
        sheet_name = f'{safe_date}_教員{route["teacher"]}'
        name_row   = route_row_map[sheet_name]
        name_ref   = f"='{INDEX_SHEET}'!C{name_row}"

        ws = wb.create_sheet(title=sheet_name)
        ws.page_setup.paperSize  = ws.PAPERSIZE_A4
        ws.page_setup.orientation = 'portrait'
        ws.page_margins.left   = 0.5
        ws.page_margins.right  = 0.5
        ws.page_margins.top    = 0.7
        ws.page_margins.bottom = 0.7

        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 24
        ws.column_dimensions['C'].width = 35
        ws.column_dimensions['D'].width = 11

        ws.row_dimensions[1].height = 34
        ws.merge_cells('A1:D1')
        ws['A1'] = f'見回り担当シート　{day["date"]}'
        ws['A1'].font = fnt(True, 14, WHITE)
        ws['A1'].fill = fill(HDR_BG)
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')

        ws.row_dimensions[2].height = 20
        ws.merge_cells('A2:D2')
        ws['A2'] = f'教員 {route["teacher"]}　担当　　※このシートを持参してください'
        ws['A2'].font = fnt(True, 10, WHITE)
        ws['A2'].fill = fill(tc)
        ws['A2'].alignment = Alignment(horizontal='center', vertical='center')

        ws.row_dimensions[3].height = 30
        ws['A3'] = '氏名'
        ws['A3'].font = fnt(True, 10, '333333')
        ws['A3'].fill = fill('EEEEEE')
        ws['A3'].border = bdr()
        ws['A3'].alignment = Alignment(horizontal='center', vertical='center')
        ws.merge_cells('B3:D3')
        ws['B3'] = name_ref
        ws['B3'].fill = fill(INPUT_BG)
        ws['B3'].border = Border(left=Side(style='medium',color='E0A800'),
                                  right=Side(style='medium',color='E0A800'),
                                  top=Side(style='medium',color='E0A800'),
                                  bottom=Side(style='medium',color='E0A800'))
        ws['B3'].alignment = Alignment(horizontal='left', vertical='center')
        ws['B3'].font = fnt(True, 12, '1A3A5C')

        ws.row_dimensions[4].height = 6

        ws.row_dimensions[5].height = 18
        for col, h in enumerate(['順','企業名　／　担当生徒','住　所　（訪問先）','距離'], 1):
            c = ws.cell(row=5, column=col, value=h)
            c.font = fnt(True, 9, WHITE)
            c.fill = fill('2D5A8E')
            c.alignment = Alignment(horizontal='center', vertical='center')
            c.border = bdr()

        r2 = 6
        for si, stop in enumerate(route['stops']):
            addr = comp_addr.get(stop['name'], '（住所未登録）')
            stu  = '　'.join(stop['students'])
            row_bg2 = WHITE if si % 2 == 0 else LIGHT

            ws.row_dimensions[r2].height   = 22
            ws.row_dimensions[r2+1].height = 18

            ws.merge_cells(f'A{r2}:A{r2+1}')
            cord = ws.cell(row=r2, column=1, value=si+1)
            cord.font = fnt(True, 13, WHITE)
            cord.fill = fill(tc)
            cord.alignment = Alignment(horizontal='center', vertical='center')
            cord.border = Border(left=Side(style='medium',color=tc),
                                  right=Side(style='thin',color='CCCCCC'),
                                  top=Side(style='medium',color=tc),
                                  bottom=Side(style='medium',color=tc))

            cn = ws.cell(row=r2, column=2, value=stop['name'])
            cn.font = fnt(True, 10, '1A3A5C')
            cn.fill = fill(row_bg2)
            cn.border = Border(left=Side(style='thin',color='CCCCCC'),
                                right=Side(style='thin',color='CCCCCC'),
                                top=Side(style='medium',color='AAAAAA'),
                                bottom=Side(style='dashed',color='CCCCCC'))
            cn.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

            cs = ws.cell(row=r2+1, column=2, value=f'担当生徒: {stu}')
            cs.font = fnt(False, 8, '555555')
            cs.fill = fill(row_bg2)
            cs.border = Border(left=Side(style='thin',color='CCCCCC'),
                                right=Side(style='thin',color='CCCCCC'),
                                top=Side(style='dashed',color='CCCCCC'),
                                bottom=Side(style='medium',color='AAAAAA'))
            cs.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

            ws.merge_cells(f'C{r2}:C{r2+1}')
            ca2 = ws.cell(row=r2, column=3, value=addr)
            ca2.font = fnt(False, 8, '333333')
            ca2.fill = fill(row_bg2)
            ca2.border = bdr()
            ca2.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

            label = '学校から' if si == 0 else '前より'
            ws.merge_cells(f'D{r2}:D{r2+1}')
            cd2 = ws.cell(row=r2, column=4, value=f'{label}\n{stop["dist_from_prev"]}km')
            cd2.font = fnt(False, 8, '666666')
            cd2.fill = fill(row_bg2)
            cd2.border = bdr()
            cd2.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

            r2 += 2

        ws.row_dimensions[r2].height = 18
        ws.merge_cells(f'A{r2}:C{r2}')
        ctot = ws.cell(row=r2, column=1,
            value=f'合計　{route["n_companies"]}社 / {route["n_students"]}名 / 計{route["total_km"]}km')
        ctot.font = fnt(True, 9, '333333')
        ctot.fill = fill('E8E8E8'); ctot.border = bdr()
        ctot.alignment = Alignment(horizontal='center', vertical='center')
        ws.cell(row=r2, column=4).fill = fill('E8E8E8')
        ws.cell(row=r2, column=4).border = bdr()

        r2 += 2
        ws.row_dimensions[r2].height = 16
        ws.merge_cells(f'A{r2}:D{r2}')
        cb2 = ws.cell(row=r2, column=1, value='備考・メモ')
        cb2.font = fnt(True, 9, WHITE); cb2.fill = fill('7F8C8D')
        cb2.alignment = Alignment(horizontal='left', vertical='center')
        cb2.border = bdr()
        r2 += 1
        for _ in range(4):
            ws.row_dimensions[r2].height = 18
            ws.merge_cells(f'A{r2}:D{r2}')
            cm = ws.cell(row=r2, column=1, value='')
            cm.fill = fill(WHITE); cm.border = bdr('thin')
            r2 += 1

wb.save(OUT)
print(f'完了: {OUT}')
