# -*- coding: utf-8 -*-
"""
教員見回り配布用.xlsx v2
- 生徒データタブを新設(生徒ID・学年・組・番号・氏名)
- 各教員タブの担当生徒欄はVLOOKUPで氏名を引っ張る形に変更
- 実際の担当教員名(260714_教員見回り配布用.xlsxより)を反映
- 距離は目安である旨の注記を追加
"""
import json, csv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE = r'C:\antigravity\public\00_作業\2026-06\インターンシップ管理'
SRC_JSON = BASE + r'\patrol_routes.json'
SRC_COMP = BASE + r'\企業マスターDB.csv'
OUT      = BASE + r'\教員見回り配布用.xlsx'

with open(SRC_JSON, encoding='utf-8') as f:
    data = json.load(f)
with open(BASE + r'\teacher_names_real.json', encoding='utf-8') as f:
    teacher_names_raw = json.load(f)
teacher_names = {}
for k, v in teacher_names_raw.items():
    date, tno = k.split('|')
    teacher_names[(date, int(tno))] = v

with open(BASE + r'\student_names.json', encoding='utf-8') as f:
    student_names = json.load(f)

comp_addr = {}
with open(SRC_COMP, encoding='utf-8-sig') as f:
    for r in csv.DictReader(f):
        comp_addr[r['企業名'].strip()] = r.get('住所', '').strip()

TEACHER_COLORS = ['C0392B','2471A3','1E8449','D68910','76448A']
HDR_BG   = '1A3A5C'
INPUT_BG = 'FFF9C4'
WHITE    = 'FFFFFF'
LIGHT    = 'F0F4FA'
NOTE_BG  = 'FDEBD0'

def fill(c): return PatternFill('solid', fgColor=c)
def fnt(bold=False, size=10, color='000000', name='Meiryo'):
    return Font(bold=bold, size=size, color=color, name=name)
def side(s='thin', c='BBBBBB'): return Side(style=s, color=c)
def bdr(t='thin'): s=side(t); return Border(left=s,right=s,top=s,bottom=s)
def gold_bdr():
    m=Side(style='medium',color='E0A800'); t=Side(style='thin',color='E0A800')
    return Border(left=m,right=m,top=t,bottom=t)

wb = Workbook()
STU_SHEET = '生徒データ'
INDEX_SHEET = '全日程一覧'

# ═══════════════ 生徒データタブ ═══════════════
ws_stu = wb.active
ws_stu.title = STU_SHEET
ws_stu.column_dimensions['A'].width = 12
ws_stu.column_dimensions['B'].width = 8
ws_stu.column_dimensions['C'].width = 8
ws_stu.column_dimensions['D'].width = 8
ws_stu.column_dimensions['E'].width = 20

ws_stu.row_dimensions[1].height = 26
ws_stu.merge_cells('A1:E1')
ws_stu['A1'] = '生徒データ（VLOOKUP参照用）'
ws_stu['A1'].font = fnt(True, 13, WHITE)
ws_stu['A1'].fill = fill(HDR_BG)
ws_stu['A1'].alignment = Alignment(horizontal='center', vertical='center')

headers_stu = ['生徒ID', '学年', '組', '番号', '氏名（ここに入力）']
for c, h in enumerate(headers_stu, 1):
    cell = ws_stu.cell(row=2, column=c, value=h)
    cell.font = fnt(True, 9, WHITE)
    cell.fill = fill('2D5A8E')
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = bdr()

all_student_ids = sorted(student_names.keys(), key=lambda k: tuple(int(x) for x in k.split('-')))
row_of_id = {}
r = 3
for sid in all_student_ids:
    nen, kumi, ban = sid.split('-')
    ws_stu.cell(row=r, column=1, value=sid).border = bdr()
    ws_stu.cell(row=r, column=1).alignment = Alignment(horizontal='center', vertical='center')
    ws_stu.cell(row=r, column=1).font = fnt(False, 9, '333333')
    ws_stu.cell(row=r, column=2, value=nen).border = bdr()
    ws_stu.cell(row=r, column=2).alignment = Alignment(horizontal='center', vertical='center')
    ws_stu.cell(row=r, column=3, value=kumi).border = bdr()
    ws_stu.cell(row=r, column=3).alignment = Alignment(horizontal='center', vertical='center')
    ws_stu.cell(row=r, column=4, value=ban).border = bdr()
    ws_stu.cell(row=r, column=4).alignment = Alignment(horizontal='center', vertical='center')
    cname = ws_stu.cell(row=r, column=5, value=student_names.get(sid, ''))
    cname.fill = fill(WHITE); cname.border = bdr()
    cname.alignment = Alignment(horizontal='left', vertical='center')
    cname.font = fnt(False, 10, '1A3A5C')
    row_of_id[sid] = r
    r += 1

STU_LAST_ROW = r - 1

def vlookup_formula(sid_cell_ref):
    return f'=IFERROR(VLOOKUP({sid_cell_ref},{STU_SHEET}!$A$3:$E${STU_LAST_ROW},5,FALSE),"")'

# ═══════════════ 全日程一覧 ═══════════════
ws_all = wb.create_sheet(INDEX_SHEET)
ws_all.column_dimensions['A'].width = 11
ws_all.column_dimensions['B'].width = 13
ws_all.column_dimensions['C'].width = 16
ws_all.column_dimensions['D'].width = 34
ws_all.column_dimensions['E'].width = 32
ws_all.column_dimensions['F'].width = 34
ws_all.column_dimensions['G'].width = 12

ws_all.row_dimensions[1].height = 30
ws_all.merge_cells('A1:G1')
ws_all['A1'] = 'インターンシップ 教員見回り担当表'
ws_all['A1'].font = fnt(True, 14, WHITE)
ws_all['A1'].fill = fill(HDR_BG)
ws_all['A1'].alignment = Alignment(horizontal='center', vertical='center')

ws_all.row_dimensions[2].height = 20
ws_all.merge_cells('A2:G2')
ws_all['A2'] = '※距離はすべて大まかな目安です。実際の道順・所要時間は各自でご確認ください。'
ws_all['A2'].font = fnt(True, 9, 'A94442')
ws_all['A2'].fill = fill(NOTE_BG)
ws_all['A2'].alignment = Alignment(horizontal='center', vertical='center')

ws_all.row_dimensions[3].height = 24
for c, h in enumerate(['日付','教員番号','担当教員名','訪問先企業','住　所','担当生徒（ID・氏名）','距離目安'], 1):
    cell = ws_all.cell(row=3, column=c, value=h)
    cell.font = fnt(True, 9, WHITE)
    cell.fill = fill('2D5A8E')
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = bdr()

route_row_map = {}
r = 4
for day in data:
    start_r = r
    for route in day['routes']:
        tno = route['teacher']
        tc = TEACHER_COLORS[(tno-1) % len(TEACHER_COLORS)]
        stops_str = ' → '.join(s['name'] for s in route['stops'])
        addr_str  = ' ／ '.join(comp_addr.get(s['name'], '') for s in route['stops'])
        all_stu   = [stu for s in route['stops'] for stu in s['students']]

        safe_date = day['date'].replace('/','-').replace('(','').replace(')','')
        sheet_name = f'{safe_date}_教員{tno}'
        route_row_map[sheet_name] = r

        ws_all.row_dimensions[r].height = 16 * max(1, len(all_stu))
        row_bg = WHITE if (r-4) % 2 == 0 else LIGHT

        ws_all.cell(row=r, column=1)

        cb = ws_all.cell(row=r, column=2, value=f'教員 {tno}')
        cb.font = fnt(True, 10, WHITE); cb.fill = fill(tc)
        cb.alignment = Alignment(horizontal='center', vertical='center')
        cb.border = bdr()

        real_name = teacher_names.get((day['date'], tno), '')
        cc = ws_all.cell(row=r, column=3, value=real_name)
        cc.fill = fill(INPUT_BG) if not real_name else fill(WHITE)
        cc.border = gold_bdr() if not real_name else bdr()
        cc.alignment = Alignment(horizontal='center', vertical='center')
        cc.font = fnt(True, 10, '1A3A5C')

        cd = ws_all.cell(row=r, column=4, value=stops_str)
        cd.fill = fill(row_bg); cd.border = bdr()
        cd.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        cd.font = fnt(False, 9)

        ce = ws_all.cell(row=r, column=5, value=addr_str)
        ce.fill = fill(row_bg); ce.border = bdr()
        ce.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        ce.font = fnt(False, 8, '555555')

        # 全日程一覧はID表示のみ。氏名確認は各教員タブでVLOOKUPにより行う
        cf = ws_all.cell(row=r, column=6, value='　'.join(all_stu))
        cf.fill = fill(row_bg); cf.border = bdr()
        cf.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        cf.font = fnt(False, 8, '333333')

        cg = ws_all.cell(row=r, column=7, value=f'約{route["total_km"]}km')
        cg.fill = fill(row_bg); cg.border = bdr()
        cg.alignment = Alignment(horizontal='center', vertical='center')
        cg.font = fnt(False, 9)

        r += 1

    end_r = r - 1
    if end_r > start_r:
        ws_all.merge_cells(f'A{start_r}:A{end_r}')
    ca = ws_all.cell(row=start_r, column=1, value=day['date'])
    ca.font = fnt(True, 10, '1A3A5C'); ca.fill = fill('D6E4F0')
    ca.alignment = Alignment(horizontal='center', vertical='center')
    ca.border = Border(left=Side(style='medium',color='1A3A5C'),
                       right=Side(style='thin',color='CCCCCC'),
                       top=Side(style='medium',color='1A3A5C'),
                       bottom=Side(style='medium',color='1A3A5C'))

# ═══════════════ 教員別詳細シート ═══════════════
import re
def dsort(d):
    m=re.search(r'(\d+)/(\d+)',d); return (int(m.group(1)),int(m.group(2))) if m else (99,99)

for day in sorted(data, key=lambda d: dsort(d['date'])):
    for route in day['routes']:
        tno = route['teacher']
        tc = TEACHER_COLORS[(tno-1) % len(TEACHER_COLORS)]
        safe_date = day['date'].replace('/','-').replace('(','').replace(')','')
        sheet_name = f'{safe_date}_教員{tno}'
        name_row   = route_row_map[sheet_name]
        name_ref   = f"='{INDEX_SHEET}'!C{name_row}"

        ws = wb.create_sheet(title=sheet_name)
        ws.page_setup.paperSize  = ws.PAPERSIZE_A4
        ws.page_setup.orientation = 'portrait'
        ws.page_margins.left   = 0.5
        ws.page_margins.right  = 0.5
        ws.page_margins.top    = 0.7
        ws.page_margins.bottom = 0.7

        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 22
        ws.column_dimensions['C'].width = 30
        ws.column_dimensions['D'].width = 10
        ws.column_dimensions['E'].width = 9
        ws.column_dimensions['F'].width = 14

        ws.row_dimensions[1].height = 32
        ws.merge_cells('A1:F1')
        ws['A1'] = f'見回り担当シート　{day["date"]}'
        ws['A1'].font = fnt(True, 14, WHITE)
        ws['A1'].fill = fill(HDR_BG)
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')

        ws.row_dimensions[2].height = 20
        ws.merge_cells('A2:F2')
        ws['A2'] = f'教員 {tno}　担当　　※このシートを持参してください'
        ws['A2'].font = fnt(True, 10, WHITE)
        ws['A2'].fill = fill(tc)
        ws['A2'].alignment = Alignment(horizontal='center', vertical='center')

        ws.row_dimensions[3].height = 28
        ws['A3'] = '氏名'
        ws['A3'].font = fnt(True, 10, '333333')
        ws['A3'].fill = fill('EEEEEE'); ws['A3'].border = bdr()
        ws['A3'].alignment = Alignment(horizontal='center', vertical='center')
        ws.merge_cells('B3:F3')
        ws['B3'] = name_ref
        ws['B3'].fill = fill(INPUT_BG)
        ws['B3'].border = Border(left=Side(style='medium',color='E0A800'),
                                  right=Side(style='medium',color='E0A800'),
                                  top=Side(style='medium',color='E0A800'),
                                  bottom=Side(style='medium',color='E0A800'))
        ws['B3'].alignment = Alignment(horizontal='left', vertical='center')
        ws['B3'].font = fnt(True, 12, '1A3A5C')

        ws.row_dimensions[4].height = 16
        ws.merge_cells('A4:F4')
        ws['A4'] = '※距離は大まかな目安です。実際の経路・所要時間は各自でご確認ください。'
        ws['A4'].font = fnt(False, 8, 'A94442')
        ws['A4'].fill = fill(NOTE_BG)
        ws['A4'].alignment = Alignment(horizontal='center', vertical='center')

        ws.row_dimensions[5].height = 18
        for col, h in enumerate(['順','企業名','住　所','距離目安','生徒ID','氏名'], 1):
            c = ws.cell(row=5, column=col, value=h)
            c.font = fnt(True, 9, WHITE)
            c.fill = fill('2D5A8E')
            c.alignment = Alignment(horizontal='center', vertical='center')
            c.border = bdr()

        r2 = 6
        for stop_idx, stop in enumerate(route['stops']):
            addr = comp_addr.get(stop['name'], '（住所未登録）')
            n_stu = max(1, len(stop['students']))
            start = r2

            for si, sid in enumerate(stop['students']):
                row_bg = WHITE if si % 2 == 0 else LIGHT
                ws.row_dimensions[r2].height = 18
                ce_id = ws.cell(row=r2, column=5, value=sid)
                ce_id.border = bdr(); ce_id.fill = fill(row_bg)
                ce_id.alignment = Alignment(horizontal='center', vertical='center')
                ce_id.font = fnt(False, 8, '333333')

                cf_name = ws.cell(row=r2, column=6, value=vlookup_formula(f'E{r2}'))
                cf_name.border = bdr(); cf_name.fill = fill(row_bg)
                cf_name.alignment = Alignment(horizontal='left', vertical='center')
                cf_name.font = fnt(False, 9, '1A3A5C')
                r2 += 1

            end = r2 - 1
            # 企業名・住所・距離・順番は結合して縦中央揃え
            ws.merge_cells(f'A{start}:A{end}')
            cord = ws.cell(row=start, column=1, value=stop_idx + 1)
            cord.font = fnt(True, 13, WHITE); cord.fill = fill(tc)
            cord.alignment = Alignment(horizontal='center', vertical='center')
            cord.border = Border(left=Side(style='medium',color=tc),
                                  right=Side(style='thin',color='CCCCCC'),
                                  top=Side(style='medium',color=tc),
                                  bottom=Side(style='medium',color=tc))

            ws.merge_cells(f'B{start}:B{end}')
            cn = ws.cell(row=start, column=2, value=stop['name'])
            cn.font = fnt(True, 10, '1A3A5C'); cn.fill = fill(WHITE)
            cn.border = bdr(); cn.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

            ws.merge_cells(f'C{start}:C{end}')
            ca2 = ws.cell(row=start, column=3, value=addr)
            ca2.font = fnt(False, 8, '333333'); ca2.fill = fill(WHITE)
            ca2.border = bdr(); ca2.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

            ws.merge_cells(f'D{start}:D{end}')
            cd2 = ws.cell(row=start, column=4, value=f'約{stop["dist_from_prev"]}km')
            cd2.font = fnt(False, 8, '666666'); cd2.fill = fill(WHITE)
            cd2.border = bdr(); cd2.alignment = Alignment(horizontal='center', vertical='center')

        # 合計行
        ws.row_dimensions[r2].height = 16
        ws.merge_cells(f'A{r2}:F{r2}')
        ctot = ws.cell(row=r2, column=1,
            value=f'合計　{route["n_companies"]}社 / {route["n_students"]}名 / 約{route["total_km"]}km')
        ctot.font = fnt(True, 9, '333333')
        ctot.fill = fill('E8E8E8'); ctot.border = bdr()
        ctot.alignment = Alignment(horizontal='center', vertical='center')
        r2 += 2

        ws.merge_cells(f'A{r2}:F{r2}')
        ws.cell(row=r2, column=1, value='備考・メモ')
        ws.cell(row=r2, column=1).font = fnt(True, 9, WHITE)
        ws.cell(row=r2, column=1).fill = fill('7F8C8D')
        ws.cell(row=r2, column=1).border = bdr()
        r2 += 1
        for _ in range(4):
            ws.row_dimensions[r2].height = 18
            ws.merge_cells(f'A{r2}:F{r2}')
            ws.cell(row=r2, column=1, value='').fill = fill(WHITE)
            ws.cell(row=r2, column=1).border = bdr('thin')
            r2 += 1

wb.save(OUT)
print(f'完了: {OUT}')
